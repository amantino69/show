# -*- coding: utf-8 -*-
"""Auditoria completa de todas as abas das planilhas."""
import json
import re
from pathlib import Path
import openpyxl

PLANILHAS = Path(__file__).resolve().parents[1] / "docs" / "planilhas"
OUT = Path(__file__).resolve().parent / "_audit_all_sheets.json"


def _str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def scan_sheet(ws, max_row=200):
    headers = []
    header_row = None
    fields = []  # label-value pairs (form style)
    data_rows = []
    sections = []

    rows = []
    for i, row in enumerate(ws.iter_rows(max_row=max_row, values_only=True), 1):
        cells = [_str(c) for c in row]
        if any(cells):
            rows.append((i, cells))

    for i, cells in rows:
        line = " | ".join(c for c in cells if c)
        upper = line.upper()

        # section headers
        if len(cells) >= 1 and cells[0].startswith("  ") and len([c for c in cells if c]) <= 2:
            sections.append(cells[0].strip())
        if "ARCHÉ" in upper or "VIEZES" in upper:
            continue
        if any(k in upper for k in ["CHECKLIST", "PIPELINE", "CRM", "ONBOARDING", "PAINEL"]):
            if len(cells) <= 3:
                continue

        # table header detection
        if not header_row and len([c for c in cells if c]) >= 4:
            hdr_keywords = ["NOME", "STATUS", "TAREFA", "CLIENTE", "MARCA", "DATA", "#", "ETAPA"]
            if sum(1 for c in cells if any(k in c.upper() for k in hdr_keywords)) >= 2:
                header_row = i
                headers = [c for c in cells if c]
                continue

        if header_row and i > header_row:
            non_empty = [c for c in cells if c and c not in ("—", "-", "☐")]
            if len(non_empty) >= 2:
                data_rows.append({"row": i, "values": cells[:15]})
            elif len(non_empty) == 1 and non_empty[0].isdigit():
                pass
        elif len(cells) >= 2 and cells[0] and cells[1] and not cells[0].isdigit():
            if cells[0] not in ("#", "STATUS", "TAREFA") and "PLATAFORMA" not in cells[0].upper():
                fields.append({"label": cells[0], "value": cells[1], "row": i})

    return {
        "headers": headers,
        "header_row": header_row,
        "sections": sections[:20],
        "form_fields": fields[:80],
        "form_field_count": len(fields),
        "data_row_count": len(data_rows),
        "data_sample": data_rows[:5],
        "total_nonempty_rows": len(rows),
    }


def main():
    report = []
    for path in sorted(PLANILHAS.glob("*.xlsx")):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        file_entry = {"file": path.name, "sheets": []}
        for sn in wb.sheetnames:
            ws = wb[sn]
            info = scan_sheet(ws)
            info["name"] = sn
            info["max_row"] = ws.max_row
            file_entry["sheets"].append(info)
        wb.close()
        report.append(file_entry)

    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Written {OUT}")


if __name__ == "__main__":
    main()
