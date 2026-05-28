# -*- coding: utf-8 -*-
import json
from pathlib import Path
import openpyxl

folder = Path(__file__).resolve().parents[1] / "docs" / "planilhas"
out = []

for f in sorted(folder.glob("*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    file_info = {"file": f.name, "sheets": []}
    for sn in wb.sheetnames:
        ws = wb[sn]
        preview = []
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            if i >= 20:
                break
            cells = [c for c in row[:15] if c is not None and str(c).strip()]
            if cells:
                preview.append([str(c) for c in cells])
        file_info["sheets"].append({"name": sn, "preview": preview})
    wb.close()
    out.append(file_info)

Path(__file__).parent.joinpath("_inspect_output.json").write_text(
    json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
)
print("written", len(out), "files")
